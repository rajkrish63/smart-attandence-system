from flask import Flask, jsonify, render_template, request, redirect, url_for, session, send_from_directory
from flask_cors import CORS
import json
import os
import subprocess
from werkzeug.utils import secure_filename
from openpyxl import load_workbook, Workbook

app = Flask(__name__)
CORS(app)
app.secret_key = os.urandom(24)

# === Fixed Folder Setup ===
BASE_DIR = os.path.abspath(os.path.dirname(__file__))

FILES_FOLDER = os.path.join(BASE_DIR, "output")
IMAGE_FOLDER = os.path.join(BASE_DIR, "Images")

# Ensure folders exist
os.makedirs(FILES_FOLDER, exist_ok=True)
os.makedirs(IMAGE_FOLDER, exist_ok=True)

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}
app.config['UPLOAD_FOLDER'] = IMAGE_FOLDER

# === Admin Credentials ===
ADMIN_USERNAME = 'admin'
ADMIN_PASSWORD = 'admin'


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# === Excel Handling Function ===
def add_to_excel(new_student):
    excel_file = "data.xlsx"

    # Create workbook if not exists
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.append(["S/No", "Roll No", "Student Name", "Parent No"])  # header
        wb.save(excel_file)

    workbook = load_workbook(excel_file)
    sheet = workbook.active

    # Get existing roll numbers
    existing_roll_numbers = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        existing_roll_numbers.add(row[1])  # Roll No is column 2

    roll_no = new_student.get("Roll No")
    if roll_no not in existing_roll_numbers:
        sheet.append([
            new_student.get("S/No"),
            new_student.get("Roll No"),
            new_student.get("Student Name"),
            new_student.get("Parent No")
        ])
        workbook.save(excel_file)


@app.route('/')
def home():
    return redirect('/login')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        if request.form['username'] == ADMIN_USERNAME and request.form['password'] == ADMIN_PASSWORD:
            session['admin_logged_in'] = True
            return redirect('/home')
        else:
            return "Invalid credentials"
    return render_template('login.html')


@app.route('/home')
def home_page():
    if not session.get('admin_logged_in'):
        return redirect('/login')
    return render_template('home.html')


@app.route('/admin')
def admin_page():
    if not session.get('admin_logged_in'):
        return redirect('/login')
    return render_template('admin.html')


@app.route('/change-time-slot', methods=['GET', 'POST'])
def change_time_slot():
    if not session.get('admin_logged_in'):
        return redirect('/login')

    if request.method == 'POST':
        start = request.form.get('start_time')
        end = request.form.get('end_time')
        if start and end:
            slots = []
            if os.path.exists('time_slots.json'):
                with open('time_slots.json', 'r') as f:
                    slots = json.load(f)
            slots.append([start, end])
            with open('time_slots.json', 'w') as f:
                json.dump(slots, f, indent=4)
            return redirect('/change-time-slot')

    current_slots = []
    if os.path.exists('time_slots.json'):
        with open('time_slots.json', 'r') as f:
            current_slots = json.load(f)

    return render_template('change_time_slot.html', slots=current_slots)


@app.route('/delete-time-slot', methods=['POST'])
def delete_time_slot():
    if not session.get('admin_logged_in'):
        return redirect('/login')

    index = int(request.form.get('index'))

    if os.path.exists('time_slots.json'):
        with open('time_slots.json', 'r') as f:
            slots = json.load(f)
        if 0 <= index < len(slots):
            slots.pop(index)
            with open('time_slots.json', 'w') as f:
                json.dump(slots, f, indent=4)

    return redirect('/change-time-slot')


@app.route('/new-student', methods=['GET', 'POST'])
def new_student():
    if not session.get('admin_logged_in'):
        return redirect('/login')

    if request.method == 'POST':
        roll_no = request.form['roll_no']
        student_name = request.form['student_name']
        parent_no = request.form['parent_no']

        if 'image' not in request.files:
            return "No image file part"

        file = request.files['image']
        if file.filename == '':
            return "No selected image"

        if file and allowed_file(file.filename):
            filename = secure_filename(f"{student_name}.jpg")
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)

            students = []
            if os.path.exists('students.json'):
                with open('students.json', 'r') as f:
                    students = json.load(f)

            new_student = {
                "S/No": len(students) + 1,
                "Roll No": roll_no,
                "Student Name": student_name,
                "Parent No": parent_no
            }

            students.append(new_student)
            with open('students.json', 'w') as f:
                json.dump(students, f, indent=4)

            # Also update Excel
            add_to_excel(new_student)

            return redirect(url_for('new_student'))

    return render_template('new_student.html')


@app.route('/files')
def list_files():
    if not session.get('admin_logged_in'):
        return {'error': 'Unauthorized'}, 401

    all_files = os.listdir(FILES_FOLDER)
    files = [f for f in all_files if f.endswith('.xlsx')]
    return jsonify(files)


@app.route('/download/<filename>')
def download_file(filename):
    if not session.get('admin_logged_in'):
        return redirect('/login')
    return send_from_directory(FILES_FOLDER, filename, as_attachment=True)


@app.route('/delete/<filename>', methods=['DELETE'])
def delete_file(filename):
    if not session.get('admin_logged_in'):
        return {'error': 'Unauthorized'}, 401

    file_path = os.path.join(FILES_FOLDER, filename)
    if os.path.exists(file_path):
        os.remove(file_path)
        return jsonify({"status": "deleted"})
    return jsonify({"error": "file not found"}), 404


@app.route('/run-python', methods=['POST'])
def run_python():
    if not session.get('admin_logged_in'):
        return jsonify({"status": "error", "output": "Unauthorized"}), 401

    try:
        result = subprocess.run(['python', 'final2.py'], capture_output=True, text=True)
        if result.returncode == 0:
            return jsonify({"status": "success", "output": result.stdout})
        else:
            return jsonify({"status": "error", "output": result.stderr}), 500
    except Exception as e:
        return jsonify({"status": "error", "output": str(e)}), 500


@app.route('/logout')
def logout():
    session.pop('admin_logged_in', None)
    return redirect('/login')


if __name__ == '__main__':
    app.run(debug=True)
