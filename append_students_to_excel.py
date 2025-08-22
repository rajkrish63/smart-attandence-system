import json
import os
from openpyxl import load_workbook, Workbook
import sys
sys.stdout.reconfigure(encoding='utf-8')

# Define absolute paths to keep consistency with app.py
BASE_DIR = "D:/auto_attendance"
JSON_FILE = os.path.join(BASE_DIR, "R:\rajkrish\git\smart-attandence-system\students.json")
EXCEL_FILE = os.path.join(BASE_DIR, "output", "data.xlsx")

def add_to_database():
    # Ensure students.json exists
    if not os.path.exists(JSON_FILE):
        print("[ERROR] students.json not found!")
        return

    with open(JSON_FILE, "r", encoding="utf-8") as f:
        students = json.load(f)

    # If Excel file doesn't exist, create it with headers
    if not os.path.exists(EXCEL_FILE):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Students"
        sheet.append(["S/No", "Roll No", "Student Name", "Parent No"])  # headers
    else:
        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook.active

    # Get all roll numbers already in Excel to prevent duplicates
    existing_roll_numbers = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[1]:  # Roll No column
            existing_roll_numbers.add(str(row[1]))

    # Append new students only if roll number doesn't exist
    for student in students:
        roll_no = str(student.get("Roll No") or student.get("roll_no"))
        if roll_no and roll_no not in existing_roll_numbers:
            sheet.append([
                student.get("S/No") or student.get("s_no"),
                roll_no,
                student.get("Student Name") or student.get("student_name"),
                student.get("Parent No") or student.get("parent_no")
            ])
            print(f" Added {roll_no} - {student.get('Student Name')}")

    # Save changes
    workbook.save(EXCEL_FILE)
    print(" Excel updated successfully at:", EXCEL_FILE)
def log_info(message, success=True):
    if success:
        print("OK:", message)
    else:
        print("ERROR:", message)
def log_error(message):
    print("ERROR:", message)

if __name__ == "__main__":
    add_to_database()
